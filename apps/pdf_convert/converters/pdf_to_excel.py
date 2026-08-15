"""
PDF -> XLSX. Real table extraction via PyMuPDF's find_tables, one sheet
per source page. Multiple tables on a page are stacked in the same sheet
with a blank row between them. A page with no detected tables still gets
its plain text written line-by-line (one line per row) rather than a
blank sheet, so nothing is silently dropped.
"""

import io

import fitz
from openpyxl import Workbook

MAX_SHEET_NAME_LENGTH = 31


def convert(file_bytes):
    """Returns (xlsx_bytes, metadata dict)."""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    try:
        workbook = Workbook()
        workbook.remove(workbook.active)  # replaced by one sheet per page below

        page_count = len(doc)
        table_count = 0

        for page_index, page in enumerate(doc):
            sheet = workbook.create_sheet(title=f"Page {page_index + 1}"[:MAX_SHEET_NAME_LENGTH])
            table_finder = page.find_tables()

            if table_finder.tables:
                row_cursor = 1
                for table in table_finder.tables:
                    for row in table.extract():
                        for col_index, value in enumerate(row, start=1):
                            sheet.cell(row=row_cursor, column=col_index, value=value)
                        row_cursor += 1
                    row_cursor += 1  # blank row between tables
                    table_count += 1
            else:
                text = page.get_text().strip()
                for row_index, line in enumerate(text.splitlines(), start=1):
                    if line.strip():
                        sheet.cell(row=row_index, column=1, value=line)

        buffer = io.BytesIO()
        workbook.save(buffer)

        metadata = {"page_count": page_count, "table_count": table_count}
        return buffer.getvalue(), metadata
    finally:
        doc.close()
